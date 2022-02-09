from unicodedata import name
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from re import I
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by  import By 
from selenium import webdriver
from openpyxl import workbook, load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

option = webdriver.ChromeOptions()
option.add_argument('headless')
driver = webdriver.Chrome("D:\\GIT\\SMC\\chromedriver.exe",options=option)
from demo1 import *
driver.get("https://shopstar.pe/tecnologia/televisores")
i=1
time.sleep(10)
 
elemnt1= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/h6[2]").text  # MODELO NOMBRE
print(elemnt1)
time.sleep(2)
elemnt2= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/span[2]/strong[1]").text # PRECIO NORMAL
print(elemnt2)
time.sleep(2)
elemnt3 = driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/div[1]").text   # PRECIO OFERTA IBK  
print(elemnt3)
time.sleep(2)
elemnt4= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[2]/a[1]").get_attribute('href')
print(elemnt4) 

print(" Modelo " + elemnt1 + " Precio Antes " +  elemnt2 + " Precio Online " + elemnt3   +  " Precio Oferta " +   elemnt4) 
'''
filepath = 'D:\GIT\shopstar\excel_data.xlsx'
wb = load_workbook(filepath)
 
def test(driver, numero_hoja):

 nombre_hoja = "Hoja1"
 ws = wb[nombre_hoja]


 driver.implicitly_wait(10)   
 time.sleep(8)   

 items = [ 1,2,3,4,5,6,7,8,9,10,11,12,13,14,115,16,17,18,19,20,21,22,23,24 ]
 
 j= 24 * numero_hoja

 for i in items :
   
   elemnt1 = "text"+str(i)
   elemnt2 = "text"+str(i)+"a"
   elemnt3 = "text"+str(i)+"b"
   elemnt4 = "text"+str(i)+"c"

   try:
       elemnt1= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+i+"]/div[1]/div[3]/h6[2]").text  # MODELO NOMBRE
       elemnt2= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+i+"]/div[1]/div[3]/span[2]/strong[1]").text # PRECIO NORMAL
       elemnt3 = driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li[variable]/div[1]/div[3]/div[1]").text   # PRECIO OFERTA IBK  
       elemnt4= driver.find_element(By.XPATH, "/html[1]/body[1]/main[1]/div[8]/section[2]/div[1]/div[2]/div[4]/div[1]/div[2]/div[2]/div[1]/ul[1]/li["+i+"]/div[1]/div[2]/div[4]/div[2]/div[2]/span[1]").get_attribute('href')
         #     elemnt4= driver.find_element(By.XPATH, "/html[1]/body[1]/main[1]/div[8]/section[2]/div[1]/div[2]/div[4]/div[1]/div[2]/div[2]/div[1]/ul[1]/li["+i+"]/div[1]/div[2]/div[4]/div[2]/div[2]/span[1]").text  
   except:
       pass
             
   print(" Modelo " + elemnt1 + " Precio Antes " +  elemnt2 + " Precio Online " + elemnt3   +  " Precio Oferta " +   elemnt4) 
 
   j +=1
  
   A="A"+str(j)
   B="B"+str(j)
   C="C"+str(j)
   D="D"+str(j)
   #ws = wb["Hoja1"]
   
   #ws = wb["Hoja"+str(numero_hoja)]
   ws[A]=elemnt1 
   #print("imprime elemnto en celda   " +A +  "  " + elemnt1  )
   ws[B]=elemnt2 
   #print("imprime elemnto en celda   " +B + "  " + elemnt2  )         
   ws[C]= elemnt3
   #print("imprime elemnto en celda   " +C+ "  " + elemnt3  )
   ws[D]=elemnt4 
   #print("imprime elemnto en celda   " +D+ "  " + elemnt4  )
   

 wb.save(filepath) 
 print("se graba exxcel")
'''