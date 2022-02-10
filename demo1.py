from unicodedata import name
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from re import I
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by  import By 
from selenium import webdriver
from openpyxl import workbook, load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait



filepath = 'D:\GIT\shopstar\excel_data.xlsx' # donde se ubica el excel
wb = load_workbook(filepath) # Abre el excel
 
def test(driver, numero_de_hoja):
 
 nombre_hoja = "Hoja1"    # variable nombre de la hoja 
 ws = wb[nombre_hoja]     # crea nombre de la hoja del excel


 driver.implicitly_wait(10)   
 time.sleep(2)   

 items = [ 1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24 ] # numero de productos en la web
 
 j= 24 * numero_de_hoja # ubica los 24 productos en las 24 filas del excel 

 for i in items :
   
   elemnt1 = "null"    # VAR QUE OPBTIENE NOMBRE DE PRODUCTO    
   elemnt2 = "S/0"     # VAR QUE OPBTIENE NOMBRE DE PRECIO DEL PRODUCTO
   elemnt = "S/0"      # VAR QUE OPBTIENE NOMBRE DE PRECIO OFERTA IBK
   elemnt4 = "S/0"     # VAR QUE OPBTIENE NOMBRE DE UNICO PRECIO DEL PRODUCTO

   try:
       elemnt1= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/h6[2]").text  # MODELO NOMBRE
       print("Modelo  "+elemnt1)      
       elemnt2= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/span[2]/strong[1]").text # PRECIO NORMAL
       print("Precio normal  "+elemnt2)      
       elemnt = driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/div[1]").text   # PRECIO OFERTA IBK  
       print("Precio IBK  "+elemnt)       
       elemnt4= driver.find_element(By.XPATH, "/html[1]/body[1]/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[2]/div[3]/div[1]/ul[1]/li["+str(i)+"]/div[1]/div[3]/span[1]").text
       print("Precio Unico   "+elemnt4) 
   except:
       pass
             
       #print(" Modelo " + elemnt1 + " Precio Normal " +  elemnt2 + " Precio IBK " + elemnt3   ) #+  " URL: " +   elemnt4) 
 
   j +=1 # VAR QUE RECORRE LAS FILAS DEL EXCEL A1,B1,C1,D1 LUGEO A2,B2,C2,D2 ......
   print( "Producto nuemro   " + j)
   A="A"+str(j) # A1,A2,A3,A4,A5....
   B="B"+str(j)  # B1,B2,B3,B4,B5....
   C="C"+str(j)   # C1,C2,C3,C4,C5....
   D="D"+str(j)     # D1,D2,D3,D4,D5....
   #ws = wb["Hoja1"]
   
   #ws = wb["Hoja"+str(numero_hoja)]
   ws[A]=elemnt1 
   #print("imprime elemnto en celda   " +A +  "  " + elemnt1  )
   ws[B]=elemnt2 
   #print("imprime elemnto en celda   " +B + "  " + elemnt2  )         
   ws[C]=elemnt
   #print("imprime elemnto en celda   " +C+ "  " + elemnt3  )
   ws[D]=elemnt4 
   #print("imprime elemnto en celda   " +D+ "  " + elemnt4  )
   #------------------------ FIN ITERACION Y REGRESA 24 VECES ( 24 PRODUCTOS ) -----------------------------
 driver.find_element(By.XPATH, "//body/div[4]/div[2]/section[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/ul[1]/li[9]/a[1]").click() # le da next a la web  al final del bucle
 time.sleep(6)
   
 
 wb.save(filepath)  # SALVA EL ARCHIVO EXCEL
 print("se graba Excel")
 
